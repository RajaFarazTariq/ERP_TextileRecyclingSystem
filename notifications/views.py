# reports/views.py
"""
Management-level reporting system.
Endpoints:
  GET /api/reports/daily-production/?date=YYYY-MM-DD
  GET /api/reports/monthly-sales/?year=YYYY&month=MM
  GET /api/reports/waste-analysis/?start=YYYY-MM-DD&end=YYYY-MM-DD
  GET /api/reports/daily-production/export/?date=YYYY-MM-DD         → Excel
  GET /api/reports/monthly-sales/export/?year=YYYY&month=MM         → Excel
  GET /api/reports/waste-analysis/export/?start=...&end=...         → Excel
"""

import io
from datetime import date, timedelta
from decimal import Decimal

from django.db.models import Sum, Count, Avg, Q
from django.http import HttpResponse
from django.utils import timezone
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

# ── Model imports (adjust app labels if your project differs) ─────────────────
from warehouse.models import Stock
from sorting.models import SortingSession, FabricStock
from decolorization.models import DecolorizationSession, ChemicalStock, ChemicalIssuance
from sales.models import SalesOrder, Payment, DispatchTracking


# ─────────────────────────────────────────────────────────────────────────────
# Safe queryset helpers — field-name-aware, works on every model
# ─────────────────────────────────────────────────────────────────────────────
def _field_names(model):
    """Return the set of concrete field names for a model."""
    return {f.name for f in model._meta.get_fields() if hasattr(f, 'column')}


def _live(queryset):
    """Filter out soft-deleted records only when the model has is_deleted."""
    if 'is_deleted' in _field_names(queryset.model):
        return queryset.filter(is_deleted=False)
    return queryset


def _date_field(model):
    """
    Return the best available date field for range filtering.
    Priority: created_at -> payment_date -> date -> updated_at
    """
    names = _field_names(model)
    for candidate in ('created_at', 'payment_date', 'date', 'updated_at'):
        if candidate in names:
            return candidate
    return None


def _date_range(queryset, start, end):
    """Filter to a date range using whatever date field the model has."""
    field = _date_field(queryset.model)
    if field:
        return queryset.filter(**{f'{field}__gte': start, f'{field}__lte': end})
    return queryset

# ─────────────────────────────────────────────────────────────────────────────
# Shared Excel helpers
# ─────────────────────────────────────────────────────────────────────────────

BRAND_GREEN  = "17703A"
BRAND_BLUE   = "1E40AF"
HEADER_FILL  = PatternFill("solid", fgColor=BRAND_BLUE)
SUBHDR_FILL  = PatternFill("solid", fgColor="E2E8F0")
ACCENT_FILL  = PatternFill("solid", fgColor="DBEAFE")
ALT_FILL     = PatternFill("solid", fgColor="F8FAFC")
WHITE_FILL   = PatternFill("solid", fgColor="FFFFFF")
HEADER_FONT  = Font(bold=True, color="FFFFFF", size=11)
SUBHDR_FONT  = Font(bold=True, color="1E293B", size=10)
TITLE_FONT   = Font(bold=True, color="0F172A", size=14)
BODY_FONT    = Font(color="334155", size=10)
MONEY_FORMAT = '#,##0'
PCT_FORMAT   = '0.0%'
DATE_FORMAT  = 'DD/MM/YYYY'

thin = Side(style="thin", color="CBD5E1")
THIN_BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)


def _col_width(ws, col, width):
    ws.column_dimensions[get_column_letter(col)].width = width


def _write_title(ws, title, subtitle, col_span):
    """Write a branded title block at the top of a sheet."""
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=col_span)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=col_span)
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=col_span)

    title_cell = ws.cell(1, 1, "Textile ERP — Recycling System")
    title_cell.font = Font(bold=True, color=BRAND_BLUE, size=16)
    title_cell.alignment = Alignment(horizontal="center")

    sub_cell = ws.cell(2, 1, title)
    sub_cell.font = TITLE_FONT
    sub_cell.alignment = Alignment(horizontal="center")

    date_cell = ws.cell(3, 1, subtitle)
    date_cell.font = Font(color="64748B", size=9, italic=True)
    date_cell.alignment = Alignment(horizontal="center")

    ws.row_dimensions[1].height = 24
    ws.row_dimensions[2].height = 20
    ws.row_dimensions[3].height = 16


def _write_header_row(ws, row, headers, fills=None):
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row, col, header)
        cell.font = HEADER_FONT
        cell.fill = fills[col - 1] if fills else HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
    ws.row_dimensions[row].height = 18


def _write_data_row(ws, row, values, money_cols=None, pct_cols=None, alt=False):
    fill = ALT_FILL if alt else WHITE_FILL
    for col, val in enumerate(values, 1):
        cell = ws.cell(row, col, val)
        cell.font = BODY_FONT
        cell.fill = fill
        cell.border = THIN_BORDER
        cell.alignment = Alignment(vertical="center")
        if money_cols and col in money_cols:
            cell.number_format = MONEY_FORMAT
            cell.alignment = Alignment(horizontal="right", vertical="center")
        if pct_cols and col in pct_cols:
            cell.number_format = '0.0"%"'


def _write_summary_block(ws, start_row, start_col, items, label_width=30):
    """Write a 2-column summary key/value block."""
    for i, (label, value, is_money) in enumerate(items):
        r = start_row + i
        lc = ws.cell(r, start_col, label)
        lc.font = Font(bold=True, color="334155", size=10)
        lc.fill = SUBHDR_FILL
        lc.border = THIN_BORDER
        lc.alignment = Alignment(vertical="center")

        vc = ws.cell(r, start_col + 1, value)
        vc.font = Font(color="1E3A8A", size=10, bold=True)
        vc.fill = ACCENT_FILL
        vc.border = THIN_BORDER
        vc.alignment = Alignment(horizontal="right", vertical="center")
        if is_money:
            vc.number_format = MONEY_FORMAT


def _int(v):
    try:
        return int(round(float(v or 0)))
    except (TypeError, ValueError):
        return 0


def _dec(v):
    try:
        return float(v or 0)
    except (TypeError, ValueError):
        return 0.0


# ─────────────────────────────────────────────────────────────────────────────
# 1. DAILY PRODUCTION REPORT
# ─────────────────────────────────────────────────────────────────────────────

def _build_daily_production_data(target_date):
    """Collect all daily production data for a given date."""
    start = timezone.make_aware(
        timezone.datetime.combine(target_date, timezone.datetime.min.time())
    )
    end = start + timedelta(days=1)

    # Warehouse stock received
    stock_received = Stock.objects.filter(
        created_at__gte=start, created_at__lt=end
    ).select_related('vendor', 'unit')

    # Sorting sessions active/completed
    sort_sessions = SortingSession.objects.filter(
        Q(created_at__gte=start, created_at__lt=end) |
        Q(end_date__gte=start, end_date__lt=end)
    ).select_related('fabric', 'supervisor')

    # Decolorization sessions
    decolor_sessions = DecolorizationSession.objects.filter(
        Q(created_at__gte=start, created_at__lt=end) |
        Q(end_date__gte=start, end_date__lt=end)
    ).select_related('tank', 'fabric', 'supervisor')

    # Summary aggregates
    stock_agg   = stock_received.aggregate(
        count=Count('id'),
        total_weight=Sum('unloading_weight')
    )
    sort_agg    = sort_sessions.aggregate(
        input=Sum('quantity_taken'),
        output=Sum('quantity_sorted'),
        waste=Sum('waste_quantity')
    )
    decolor_agg = decolor_sessions.aggregate(
        input=Sum('input_quantity'),
        output=Sum('output_quantity'),
        waste=Sum('waste_quantity')
    )

    return {
        'date': target_date,
        'stock_received': stock_received,
        'sort_sessions':  sort_sessions,
        'decolor_sessions': decolor_sessions,
        'stock_agg':   stock_agg,
        'sort_agg':    sort_agg,
        'decolor_agg': decolor_agg,
    }


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def daily_production_report(request):
    """JSON endpoint for daily production data."""
    date_str = request.GET.get('date', str(date.today()))
    try:
        target_date = date.fromisoformat(date_str)
    except ValueError:
        return Response({'error': 'Invalid date format. Use YYYY-MM-DD.'}, status=400)

    d = _build_daily_production_data(target_date)

    return Response({
        'date': str(d['date']),
        'warehouse': {
            'stock_entries': d['stock_agg']['count'] or 0,
            'total_weight_kg': _int(d['stock_agg']['total_weight']),
        },
        'sorting': {
            'sessions': d['sort_sessions'].count(),
            'input_kg': _int(d['sort_agg']['input']),
            'output_kg': _int(d['sort_agg']['output']),
            'waste_kg': _int(d['sort_agg']['waste']),
            'efficiency_pct': round(
                (_int(d['sort_agg']['output']) / _int(d['sort_agg']['input']) * 100)
                if _int(d['sort_agg']['input']) > 0 else 0, 1
            ),
        },
        'decolorization': {
            'sessions': d['decolor_sessions'].count(),
            'input_kg': _int(d['decolor_agg']['input']),
            'output_kg': _int(d['decolor_agg']['output']),
            'waste_kg': _int(d['decolor_agg']['waste']),
            'efficiency_pct': round(
                (_int(d['decolor_agg']['output']) / _int(d['decolor_agg']['input']) * 100)
                if _int(d['decolor_agg']['input']) > 0 else 0, 1
            ),
        },
    })


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def daily_production_export(request):
    """Export daily production report as Excel."""
    date_str = request.GET.get('date', str(date.today()))
    try:
        target_date = date.fromisoformat(date_str)
    except ValueError:
        return HttpResponse('Invalid date', status=400)

    d = _build_daily_production_data(target_date)
    wb = openpyxl.Workbook()

    # ── Sheet 1: Summary ──────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Summary"
    _write_title(ws, "Daily Production Report",
                 f"Date: {target_date.strftime('%A, %d %B %Y')} | Generated: {date.today().strftime('%d/%m/%Y')}", 4)

    # Summary block
    sort_in  = _int(d['sort_agg']['input'])
    sort_out = _int(d['sort_agg']['output'])
    sort_wst = _int(d['sort_agg']['waste'])
    dec_in   = _int(d['decolor_agg']['input'])
    dec_out  = _int(d['decolor_agg']['output'])
    dec_wst  = _int(d['decolor_agg']['waste'])

    ws.cell(5, 1, "PRODUCTION SUMMARY").font = Font(bold=True, color=BRAND_BLUE, size=12)
    summary_items = [
        ("Warehouse Stock Entries",     d['stock_agg']['count'] or 0,  False),
        ("Warehouse Weight Received",   _int(d['stock_agg']['total_weight']), False),
        ("Sorting Sessions",            d['sort_sessions'].count(),     False),
        ("Sorting Input (kg)",          sort_in,                        False),
        ("Sorting Output (kg)",         sort_out,                       False),
        ("Sorting Waste (kg)",          sort_wst,                       False),
        ("Sorting Efficiency (%)",      f"{round(sort_out/sort_in*100,1) if sort_in else 0}%", False),
        ("Decolorization Sessions",     d['decolor_sessions'].count(),  False),
        ("Decolorization Input (kg)",   dec_in,                         False),
        ("Decolorization Output (kg)",  dec_out,                        False),
        ("Decolorization Waste (kg)",   dec_wst,                        False),
        ("Decolorization Efficiency",   f"{round(dec_out/dec_in*100,1) if dec_in else 0}%", False),
    ]
    _write_summary_block(ws, 6, 1, summary_items)

    _col_width(ws, 1, 34)
    _col_width(ws, 2, 20)

    # ── Sheet 2: Warehouse Receipts ───────────────────────────────────────────
    ws2 = wb.create_sheet("Warehouse Receipts")
    _write_title(ws2, "Warehouse Receipts", f"Date: {target_date}", 7)
    headers = ["#", "Vendor", "Fabric Type", "Slip No.", "Vehicle", "Weight (kg)", "Unit"]
    _write_header_row(ws2, 5, headers)
    for i, s in enumerate(d['stock_received'], 1):
        _write_data_row(ws2, 5 + i, [
            i, s.vendor.name, s.fabric_type, s.vendor_weight_slip,
            s.vehicle_no, _int(s.unloading_weight), s.unit.name if s.unit else '—'
        ], alt=(i % 2 == 0))
    for col, w in zip(range(1, 8), [4, 22, 18, 14, 12, 12, 10]):
        _col_width(ws2, col, w)

    # ── Sheet 3: Sorting Sessions ─────────────────────────────────────────────
    ws3 = wb.create_sheet("Sorting Sessions")
    _write_title(ws3, "Sorting Sessions", f"Date: {target_date}", 8)
    headers = ["#", "Fabric", "Supervisor", "Unit", "Input (kg)", "Output (kg)", "Waste (kg)", "Efficiency"]
    _write_header_row(ws3, 5, headers)
    for i, s in enumerate(d['sort_sessions'], 1):
        inp = _int(s.quantity_taken)
        out = _int(s.quantity_sorted)
        eff = f"{round(out/inp*100,1)}%" if inp > 0 else "—"
        _write_data_row(ws3, 5 + i, [
            i, s.fabric.material_type if s.fabric else '—',
            s.supervisor.username if s.supervisor else '—',
            s.unit, inp, out, _int(s.waste_quantity), eff
        ], money_cols={5, 6, 7}, alt=(i % 2 == 0))
    for col, w in zip(range(1, 9), [4, 26, 16, 10, 12, 12, 12, 12]):
        _col_width(ws3, col, w)

    # ── Sheet 4: Decolorization Sessions ─────────────────────────────────────
    ws4 = wb.create_sheet("Decolorization Sessions")
    _write_title(ws4, "Decolorization Sessions", f"Date: {target_date}", 8)
    headers = ["#", "Tank", "Fabric", "Supervisor", "Input (kg)", "Output (kg)", "Waste (kg)", "Efficiency"]
    _write_header_row(ws4, 5, headers)
    for i, s in enumerate(d['decolor_sessions'], 1):
        inp = _int(s.input_quantity)
        out = _int(s.output_quantity)
        eff = f"{round(out/inp*100,1)}%" if inp > 0 else "—"
        _write_data_row(ws4, 5 + i, [
            i, s.tank.name if s.tank else '—',
            s.fabric.material_type if s.fabric else '—',
            s.supervisor.username if s.supervisor else '—',
            inp, out, _int(s.waste_quantity), eff
        ], money_cols={5, 6, 7}, alt=(i % 2 == 0))
    for col, w in zip(range(1, 9), [4, 12, 26, 16, 12, 12, 12, 12]):
        _col_width(ws4, col, w)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"daily_production_{target_date}.xlsx"
    response = HttpResponse(
        buf.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


# ─────────────────────────────────────────────────────────────────────────────
# 2. MONTHLY SALES REPORT
# ─────────────────────────────────────────────────────────────────────────────

def _build_monthly_sales_data(year, month):
    from calendar import monthrange
    _, last_day = monthrange(year, month)
    start = timezone.make_aware(timezone.datetime(year, month, 1))
    end   = timezone.make_aware(timezone.datetime(year, month, last_day, 23, 59, 59))

    orders   = _live(_date_range(SalesOrder.objects.all(),       start, end)).select_related('fabric', 'created_by')
    payments = _live(_date_range(Payment.objects.all(),          start, end)).select_related('sales_order', 'received_by')
    dispatch = _live(_date_range(DispatchTracking.objects.all(), start, end)).select_related('sales_order')

    order_agg = orders.aggregate(
        count=Count('id'),
        total_revenue=Sum('total_price'),
        total_weight=Sum('weight_sold'),
        avg_price=Avg('price_per_kg'),
    )
    payment_agg = payments.aggregate(
        count=Count('id'),
        total_collected=Sum('amount'),
    )

    return {
        'year': year, 'month': month,
        'orders': orders, 'payments': payments, 'dispatch': dispatch,
        'order_agg': order_agg, 'payment_agg': payment_agg,
    }


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def monthly_sales_report(request):
    today = date.today()
    try:
        year  = int(request.GET.get('year',  today.year))
        month = int(request.GET.get('month', today.month))
    except ValueError:
        return Response({'error': 'Invalid year/month.'}, status=400)

    d = _build_monthly_sales_data(year, month)
    orders = d['orders']

    # Status breakdown
    status_breakdown = {}
    for o in orders:
        st = o.status or 'Unknown'
        status_breakdown[st] = status_breakdown.get(st, 0) + 1

    # Payment breakdown
    pay_breakdown = {}
    for p in d['payments']:
        m = p.payment_method or 'Other'
        if m not in pay_breakdown:
            pay_breakdown[m] = {'count': 0, 'amount': 0}
        pay_breakdown[m]['count']  += 1
        pay_breakdown[m]['amount'] += _int(p.amount)

    total_rev = _int(d['order_agg']['total_revenue'])
    total_col = _int(d['payment_agg']['total_collected'])

    return Response({
        'year': year, 'month': month,
        'total_orders':    d['order_agg']['count'] or 0,
        'total_revenue':   total_rev,
        'total_collected': total_col,
        'pending_amount':  max(0, total_rev - total_col),
        'total_weight_kg': _int(d['order_agg']['total_weight']),
        'avg_price_per_kg': round(_dec(d['order_agg']['avg_price']), 0),
        'total_dispatches': d['dispatch'].count(),
        'status_breakdown': status_breakdown,
        'payment_method_breakdown': pay_breakdown,
    })


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def monthly_sales_export(request):
    today = date.today()
    try:
        year  = int(request.GET.get('year',  today.year))
        month = int(request.GET.get('month', today.month))
    except ValueError:
        return HttpResponse('Invalid year/month', status=400)

    import calendar
    d = _build_monthly_sales_data(year, month)
    month_name = calendar.month_name[month]
    wb = openpyxl.Workbook()

    # ── Sheet 1: Summary ──────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Summary"
    _write_title(ws, f"Monthly Sales Report — {month_name} {year}",
                 f"Generated: {date.today().strftime('%d/%m/%Y')}", 4)

    total_rev = _int(d['order_agg']['total_revenue'])
    total_col = _int(d['payment_agg']['total_collected'])
    summary_items = [
        ("Total Orders",          d['order_agg']['count'] or 0,            False),
        ("Total Revenue (Rs.)",   total_rev,                                True),
        ("Amount Collected (Rs.)", total_col,                               True),
        ("Pending Amount (Rs.)",  max(0, total_rev - total_col),            True),
        ("Total Weight Sold (kg)", _int(d['order_agg']['total_weight']),    False),
        ("Avg Price / kg (Rs.)",   round(_dec(d['order_agg']['avg_price'])), False),
        ("Total Payments",        d['payment_agg']['count'] or 0,           False),
        ("Total Dispatches",      d['dispatch'].count(),                    False),
    ]
    ws.cell(5, 1, "MONTHLY SALES SUMMARY").font = Font(bold=True, color=BRAND_BLUE, size=12)
    _write_summary_block(ws, 6, 1, summary_items)
    _col_width(ws, 1, 30)
    _col_width(ws, 2, 20)

    # ── Sheet 2: All Orders ───────────────────────────────────────────────────
    ws2 = wb.create_sheet("Sales Orders")
    _write_title(ws2, f"Sales Orders — {month_name} {year}", "", 9)
    headers = ["#", "Order ID", "Buyer", "Contact", "Quality", "Weight (kg)", "Price/kg", "Total (Rs.)", "Payment Status"]
    _write_header_row(ws2, 5, headers)
    for i, o in enumerate(d['orders'], 1):
        _write_data_row(ws2, 5 + i, [
            i, f"#{o.id}", o.buyer_name, o.buyer_contact or '—',
            o.fabric_quality or '—', _int(o.weight_sold),
            _int(o.price_per_kg), _int(o.total_price), o.payment_status
        ], money_cols={7, 8}, alt=(i % 2 == 0))
    for col, w in zip(range(1, 10), [4, 10, 26, 16, 12, 12, 12, 16, 14]):
        _col_width(ws2, col, w)

    # ── Sheet 3: Payments ─────────────────────────────────────────────────────
    ws3 = wb.create_sheet("Payments")
    _write_title(ws3, f"Payments — {month_name} {year}", "", 6)
    headers = ["#", "Order", "Buyer", "Amount (Rs.)", "Method", "Reference"]
    _write_header_row(ws3, 5, headers)
    for i, p in enumerate(d['payments'], 1):
        buyer = p.sales_order.buyer_name if p.sales_order else '—'
        _write_data_row(ws3, 5 + i, [
            i, f"#{p.sales_order_id}", buyer, _int(p.amount),
            p.payment_method, p.reference_number or '—'
        ], money_cols={4}, alt=(i % 2 == 0))
    for col, w in zip(range(1, 7), [4, 10, 26, 16, 16, 18]):
        _col_width(ws3, col, w)

    # ── Sheet 4: Dispatch ─────────────────────────────────────────────────────
    ws4 = wb.create_sheet("Dispatch")
    _write_title(ws4, f"Dispatch — {month_name} {year}", "", 6)
    headers = ["#", "Order", "Vehicle", "Driver", "Weight (kg)", "Status"]
    _write_header_row(ws4, 5, headers)
    for i, dp in enumerate(d['dispatch'], 1):
        _write_data_row(ws4, 5 + i, [
            i, f"#{dp.sales_order_id}", dp.vehicle_number,
            dp.driver_name, _int(dp.dispatched_weight), dp.dispatch_status
        ], alt=(i % 2 == 0))
    for col, w in zip(range(1, 7), [4, 10, 14, 20, 14, 14]):
        _col_width(ws4, col, w)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"monthly_sales_{year}_{month:02d}.xlsx"
    response = HttpResponse(
        buf.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


# ─────────────────────────────────────────────────────────────────────────────
# 3. WASTE ANALYSIS REPORT
# ─────────────────────────────────────────────────────────────────────────────

def _build_waste_data(start_date, end_date):
    start = timezone.make_aware(timezone.datetime.combine(start_date, timezone.datetime.min.time()))
    end   = timezone.make_aware(timezone.datetime.combine(end_date,   timezone.datetime.max.time()))

    sort_sessions   = SortingSession.objects.filter(
        status='Completed',
        end_date__gte=start, end_date__lte=end
    ).select_related('fabric', 'supervisor')

    decolor_sessions = DecolorizationSession.objects.filter(
        status='Completed',
        end_date__gte=start, end_date__lte=end
    ).select_related('tank', 'fabric', 'supervisor')

    sort_agg    = sort_sessions.aggregate(input=Sum('quantity_taken'), waste=Sum('waste_quantity'), output=Sum('quantity_sorted'))
    decolor_agg = decolor_sessions.aggregate(input=Sum('input_quantity'), waste=Sum('waste_quantity'), output=Sum('output_quantity'))

    # Per-fabric waste breakdown (sorting)
    fabric_waste = {}
    for s in sort_sessions:
        name = s.fabric.material_type if s.fabric else 'Unknown'
        if name not in fabric_waste:
            fabric_waste[name] = {'input': 0, 'output': 0, 'waste': 0}
        fabric_waste[name]['input']  += _int(s.quantity_taken)
        fabric_waste[name]['output'] += _int(s.quantity_sorted)
        fabric_waste[name]['waste']  += _int(s.waste_quantity)

    return {
        'start': start_date, 'end': end_date,
        'sort_sessions': sort_sessions,
        'decolor_sessions': decolor_sessions,
        'sort_agg': sort_agg,
        'decolor_agg': decolor_agg,
        'fabric_waste': fabric_waste,
    }


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def waste_analysis_report(request):
    today = date.today()
    try:
        start_date = date.fromisoformat(request.GET.get('start', str(today - timedelta(days=30))))
        end_date   = date.fromisoformat(request.GET.get('end',   str(today)))
    except ValueError:
        return Response({'error': 'Invalid date format.'}, status=400)

    d = _build_waste_data(start_date, end_date)

    sort_in  = _int(d['sort_agg']['input'])
    sort_wst = _int(d['sort_agg']['waste'])
    dec_in   = _int(d['decolor_agg']['input'])
    dec_wst  = _int(d['decolor_agg']['waste'])
    total_in  = sort_in + dec_in
    total_wst = sort_wst + dec_wst

    return Response({
        'start': str(start_date), 'end': str(end_date),
        'sorting': {
            'sessions':    d['sort_sessions'].count(),
            'input_kg':    sort_in,
            'waste_kg':    sort_wst,
            'waste_pct':   round(sort_wst / sort_in * 100, 1) if sort_in else 0,
        },
        'decolorization': {
            'sessions':    d['decolor_sessions'].count(),
            'input_kg':    dec_in,
            'waste_kg':    dec_wst,
            'waste_pct':   round(dec_wst / dec_in * 100, 1) if dec_in else 0,
        },
        'total': {
            'input_kg':    total_in,
            'waste_kg':    total_wst,
            'waste_pct':   round(total_wst / total_in * 100, 1) if total_in else 0,
        },
        'by_fabric': [
            {
                'fabric': name,
                'input_kg': v['input'],
                'output_kg': v['output'],
                'waste_kg': v['waste'],
                'waste_pct': round(v['waste'] / v['input'] * 100, 1) if v['input'] else 0,
            }
            for name, v in sorted(d['fabric_waste'].items(), key=lambda x: -x[1]['waste'])
        ],
    })


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def waste_analysis_export(request):
    today = date.today()
    try:
        start_date = date.fromisoformat(request.GET.get('start', str(today - timedelta(days=30))))
        end_date   = date.fromisoformat(request.GET.get('end',   str(today)))
    except ValueError:
        return HttpResponse('Invalid dates', status=400)

    d = _build_waste_data(start_date, end_date)
    wb = openpyxl.Workbook()

    # ── Sheet 1: Summary ──────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Waste Summary"
    period = f"{start_date.strftime('%d/%m/%Y')} to {end_date.strftime('%d/%m/%Y')}"
    _write_title(ws, "Waste Analysis Report", f"Period: {period} | Generated: {date.today().strftime('%d/%m/%Y')}", 4)

    sort_in  = _int(d['sort_agg']['input'])
    sort_wst = _int(d['sort_agg']['waste'])
    dec_in   = _int(d['decolor_agg']['input'])
    dec_wst  = _int(d['decolor_agg']['waste'])
    total_in  = sort_in + dec_in
    total_wst = sort_wst + dec_wst

    ws.cell(5, 1, "WASTE ANALYSIS SUMMARY").font = Font(bold=True, color=BRAND_BLUE, size=12)
    summary_items = [
        ("Sorting Sessions Analysed",       d['sort_sessions'].count(),  False),
        ("Sorting — Total Input (kg)",       sort_in,                     False),
        ("Sorting — Total Waste (kg)",       sort_wst,                    False),
        ("Sorting — Waste Rate",             f"{round(sort_wst/sort_in*100,1) if sort_in else 0}%", False),
        ("Decolorization Sessions Analysed", d['decolor_sessions'].count(), False),
        ("Decolorization — Total Input (kg)", dec_in,                    False),
        ("Decolorization — Total Waste (kg)", dec_wst,                   False),
        ("Decolorization — Waste Rate",       f"{round(dec_wst/dec_in*100,1) if dec_in else 0}%", False),
        ("OVERALL Total Input (kg)",          total_in,                  False),
        ("OVERALL Total Waste (kg)",          total_wst,                 False),
        ("OVERALL Waste Rate",                f"{round(total_wst/total_in*100,1) if total_in else 0}%", False),
    ]
    _write_summary_block(ws, 6, 1, summary_items)
    _col_width(ws, 1, 36)
    _col_width(ws, 2, 20)

    # ── Sheet 2: By Fabric ────────────────────────────────────────────────────
    ws2 = wb.create_sheet("Waste by Fabric")
    _write_title(ws2, "Waste by Fabric Type", f"Period: {period}", 6)
    headers = ["#", "Fabric / Material", "Input (kg)", "Output (kg)", "Waste (kg)", "Waste Rate (%)"]
    _write_header_row(ws2, 5, headers)
    for i, (name, v) in enumerate(
        sorted(d['fabric_waste'].items(), key=lambda x: -x[1]['waste']), 1
    ):
        wst_pct = f"{round(v['waste']/v['input']*100,1)}%" if v['input'] else "—"
        _write_data_row(ws2, 5 + i, [
            i, name, v['input'], v['output'], v['waste'], wst_pct
        ], alt=(i % 2 == 0))
    for col, w in zip(range(1, 7), [4, 30, 14, 14, 14, 14]):
        _col_width(ws2, col, w)

    # ── Sheet 3: Sorting Detail ───────────────────────────────────────────────
    ws3 = wb.create_sheet("Sorting Detail")
    _write_title(ws3, "Sorting Waste Detail", f"Period: {period}", 7)
    headers = ["#", "Fabric", "Supervisor", "Unit", "Input (kg)", "Waste (kg)", "Waste Rate"]
    _write_header_row(ws3, 5, headers)
    for i, s in enumerate(d['sort_sessions'], 1):
        inp = _int(s.quantity_taken)
        wst = _int(s.waste_quantity)
        _write_data_row(ws3, 5 + i, [
            i, s.fabric.material_type if s.fabric else '—',
            s.supervisor.username if s.supervisor else '—',
            s.unit, inp, wst,
            f"{round(wst/inp*100,1)}%" if inp else "—"
        ], alt=(i % 2 == 0))
    for col, w in zip(range(1, 8), [4, 26, 16, 10, 12, 12, 12]):
        _col_width(ws3, col, w)

    # ── Sheet 4: Decolorization Detail ───────────────────────────────────────
    ws4 = wb.create_sheet("Decolorization Detail")
    _write_title(ws4, "Decolorization Waste Detail", f"Period: {period}", 7)
    headers = ["#", "Tank", "Fabric", "Supervisor", "Input (kg)", "Waste (kg)", "Waste Rate"]
    _write_header_row(ws4, 5, headers)
    for i, s in enumerate(d['decolor_sessions'], 1):
        inp = _int(s.input_quantity)
        wst = _int(s.waste_quantity)
        _write_data_row(ws4, 5 + i, [
            i, s.tank.name if s.tank else '—',
            s.fabric.material_type if s.fabric else '—',
            s.supervisor.username if s.supervisor else '—',
            inp, wst,
            f"{round(wst/inp*100,1)}%" if inp else "—"
        ], alt=(i % 2 == 0))
    for col, w in zip(range(1, 8), [4, 14, 26, 16, 12, 12, 12]):
        _col_width(ws4, col, w)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"waste_analysis_{start_date}_to_{end_date}.xlsx"
    response = HttpResponse(
        buf.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response