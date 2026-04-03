# notifications/tasks.py
"""
Email alert system for Textile ERP.
Scheduled reports send HTML email + Excel file attached.
─────────────────────────────────────────────────────────────
SETTINGS REQUIRED in settings.py:
  EMAIL_BACKEND       = 'django.core.mail.backends.smtp.EmailBackend'
  EMAIL_HOST          = 'smtp.gmail.com'
  EMAIL_PORT          = 587
  EMAIL_USE_TLS       = True
  EMAIL_HOST_USER     = 'your-gmail@gmail.com'
  EMAIL_HOST_PASSWORD = 'xxxx xxxx xxxx xxxx'   # Gmail App Password
  DEFAULT_FROM_EMAIL  = 'Textile ERP <your-gmail@gmail.com>'
  MANAGEMENT_EMAIL    = 'thisismefaraz@gmail.com'
─────────────────────────────────────────────────────────────
"""

import io
import logging
import calendar
from datetime import date, timedelta

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from django.conf import settings
from django.core.mail import send_mail, EmailMultiAlternatives
from django.db.models import Sum, Count, Avg
from django.utils import timezone

logger = logging.getLogger(__name__)

MANAGEMENT_EMAIL = getattr(settings, 'MANAGEMENT_EMAIL', 'thisismefaraz@gmail.com')
FROM_EMAIL       = getattr(settings, 'DEFAULT_FROM_EMAIL', 'Textile ERP <noreply@textile-erp.com>')

CHEMICAL_LOW_PCT = 0.25
WAREHOUSE_LOW_KG = 1000


# ─────────────────────────────────────────────────────────────────────────────
# Safe queryset helpers — field-name-aware, works on every model
# ─────────────────────────────────────────────────────────────────────────────

def _field_names(model):
    return {f.name for f in model._meta.get_fields() if hasattr(f, 'column')}


def _live(queryset):
    """Filter out soft-deleted records only when the model has is_deleted."""
    if 'is_deleted' in _field_names(queryset.model):
        return queryset.filter(is_deleted=False)
    return queryset


def _date_field(model):
    """
    Detect which date field a model uses.
    Priority: created_at → payment_date → date → updated_at
    """
    names = _field_names(model)
    for candidate in ('created_at', 'start_date', 'payment_date', 'date', 'updated_at'):
        if candidate in names:
            return candidate
    return None


def _date_range(queryset, start, end):
    """Filter by date range using whatever date field the model has."""
    field = _date_field(queryset.model)
    if field:
        return queryset.filter(**{f'{field}__gte': start, f'{field}__lte': end})
    return queryset


def _int(v):
    try:
        return int(round(float(v or 0)))
    except (TypeError, ValueError):
        return 0


# ─────────────────────────────────────────────────────────────────────────────
# Shared Excel styling helpers
# ─────────────────────────────────────────────────────────────────────────────

_BLUE_HEX  = "1E40AF"
_GREEN_HEX = "059669"
_thin_side = Side(style="thin", color="CBD5E1")
_BORDER    = Border(
    left=_thin_side, right=_thin_side,
    top=_thin_side,  bottom=_thin_side,
)
_HDR_FILL   = PatternFill("solid", fgColor=_BLUE_HEX)
_ALT_FILL   = PatternFill("solid", fgColor="F1F5F9")
_WHITE_FILL = PatternFill("solid", fgColor="FFFFFF")
_SUB_FILL   = PatternFill("solid", fgColor="E2E8F0")
_ACC_FILL   = PatternFill("solid", fgColor="DBEAFE")


def _cw(ws, col, width):
    ws.column_dimensions[get_column_letter(col)].width = width


def _xl_title(ws, line1, line2, line3, span):
    """Write a 3-row branded title block."""
    for r in range(1, 4):
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=span)
    ws.cell(1, 1, line1).font  = Font(bold=True, color=_BLUE_HEX, size=15)
    ws.cell(2, 1, line2).font  = Font(bold=True, color="0F172A",  size=13)
    ws.cell(3, 1, line3).font  = Font(italic=True, color="64748B", size=9)
    for r in range(1, 4):
        ws.cell(r, 1).alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 18
    ws.row_dimensions[3].height = 14


def _xl_header(ws, row, headers):
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row, c, h)
        cell.font      = Font(bold=True, color="FFFFFF", size=10)
        cell.fill      = _HDR_FILL
        cell.border    = _BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[row].height = 16


def _xl_row(ws, row, values, money_cols=None, alt=False):
    fill = _ALT_FILL if alt else _WHITE_FILL
    for c, v in enumerate(values, 1):
        cell = ws.cell(row, c, v)
        cell.font      = Font(color="334155", size=10)
        cell.fill      = fill
        cell.border    = _BORDER
        cell.alignment = Alignment(vertical="center")
        if money_cols and c in money_cols:
            cell.number_format = '#,##0'
            cell.alignment     = Alignment(horizontal="right", vertical="center")


def _xl_summary(ws, start_row, col, items):
    """Write a label/value summary block."""
    for i, (label, value) in enumerate(items):
        r = start_row + i
        lc = ws.cell(r, col, label)
        lc.font = Font(bold=True, color="334155", size=10)
        lc.fill = _SUB_FILL
        lc.border = _BORDER
        vc = ws.cell(r, col + 1, value)
        vc.font = Font(bold=True, color="1E3A8A", size=10)
        vc.fill = _ACC_FILL
        vc.border = _BORDER
        vc.alignment = Alignment(horizontal="right")


def _wb_bytes(wb):
    """Serialize a Workbook to bytes."""
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ─────────────────────────────────────────────────────────────────────────────
# Email senders
# ─────────────────────────────────────────────────────────────────────────────

def _send_alert(subject, text_body, html_body=None):
    """Send a plain or HTML email to management. Logs errors, does not raise."""
    try:
        if html_body:
            msg = EmailMultiAlternatives(subject, text_body, FROM_EMAIL, [MANAGEMENT_EMAIL])
            msg.attach_alternative(html_body, "text/html")
            msg.send(fail_silently=False)
        else:
            send_mail(subject, text_body, FROM_EMAIL, [MANAGEMENT_EMAIL], fail_silently=False)
        logger.info(f"Alert sent: {subject}")
    except Exception as exc:
        logger.error(f"Alert failed '{subject}': {exc}")
        raise


def _send_with_attachment(subject, text_body, html_body, xl_bytes, xl_filename):
    """
    Send an HTML email with an Excel (.xlsx) file attached.
    Used for all scheduled reports (monthly, daily).
    """
    msg = EmailMultiAlternatives(subject, text_body, FROM_EMAIL, [MANAGEMENT_EMAIL])
    msg.attach_alternative(html_body, "text/html")
    msg.attach(
        xl_filename,
        xl_bytes,
        'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    msg.send(fail_silently=False)
    logger.info(f"Report email sent with attachment '{xl_filename}'")


def _html_wrap(title, body_rows, color="#1E40AF", note="Excel report attached."):
    rows_html = "".join(
        f'<tr>'
        f'<td style="padding:7px 14px;border-bottom:1px solid #e2e8f0;color:#334155;font-size:13px;">'
        f'{label}</td>'
        f'<td style="padding:7px 14px;border-bottom:1px solid #e2e8f0;font-weight:600;'
        f'color:#0f172a;text-align:right;font-size:13px;">{value}</td>'
        f'</tr>'
        for label, value in body_rows
    )
    return f"""
    <div style="font-family:Arial,sans-serif;max-width:560px;margin:0 auto;
                background:#f8fafc;padding:24px;border-radius:12px;">
      <div style="background:{color};border-radius:8px 8px 0 0;padding:18px 24px;">
        <h2 style="color:white;margin:0;font-size:18px;letter-spacing:-0.3px;">Textile ERP</h2>
        <p style="color:rgba(255,255,255,0.85);margin:4px 0 0;font-size:13px;">{title}</p>
      </div>
      <div style="background:white;border-radius:0 0 8px 8px;">
        <table style="width:100%;border-collapse:collapse;">{rows_html}</table>
      </div>
      <p style="color:#94a3b8;font-size:11px;text-align:center;margin-top:12px;">
        {note} &middot; Automated alert &middot; Textile ERP &middot;
        {date.today().strftime('%d %b %Y')}
      </p>
    </div>
    """


# ─────────────────────────────────────────────────────────────────────────────
# 1. ORDER ALERTS
# ─────────────────────────────────────────────────────────────────────────────

def alert_order_completed(order):
    rows = [
        ("Order ID",       f"#{order.id}"),
        ("Buyer",          order.buyer_name),
        ("Fabric Quality", order.fabric_quality or "—"),
        ("Weight Sold",    f"{_int(order.weight_sold):,} kg"),
        ("Total Amount",   f"Rs. {_int(order.total_price):,}"),
        ("Payment Status", order.payment_status),
        ("Date",           date.today().strftime('%d %b %Y')),
    ]
    _send_alert(
        subject=f"[ERP] Order #{order.id} Completed — {order.buyer_name}",
        text_body=f"ORDER COMPLETED\nOrder #{order.id} — {order.buyer_name}\nAmount: Rs. {_int(order.total_price):,}",
        html_body=_html_wrap("Order Completed", rows, color="#059669", note=""),
    )


def alert_order_dispatched(order, dispatch):
    rows = [
        ("Order ID",          f"#{order.id}"),
        ("Buyer",             order.buyer_name),
        ("Vehicle",           dispatch.vehicle_number),
        ("Driver",            dispatch.driver_name),
        ("Driver Contact",    dispatch.driver_contact or "—"),
        ("Weight Dispatched", f"{_int(dispatch.dispatched_weight):,} kg"),
        ("Date",              date.today().strftime('%d %b %Y')),
    ]
    _send_alert(
        subject=f"[ERP] Order #{order.id} Dispatched — {order.buyer_name}",
        text_body=f"ORDER DISPATCHED\nOrder #{order.id} — {order.buyer_name}\nVehicle: {dispatch.vehicle_number}",
        html_body=_html_wrap("Order Dispatched", rows, color="#7C3AED", note=""),
    )


# ─────────────────────────────────────────────────────────────────────────────
# 2. PAYMENT ALERTS
# ─────────────────────────────────────────────────────────────────────────────

def alert_payment_received(payment):
    order = payment.sales_order
    buyer = order.buyer_name if order else "—"
    rows = [
        ("Payment ID",      f"#{payment.id}"),
        ("Order",           f"#{order.id} — {buyer}" if order else "—"),
        ("Amount Received", f"Rs. {_int(payment.amount):,}"),
        ("Method",          payment.payment_method),
        ("Reference",       payment.reference_number or "—"),
        ("Received By",     payment.received_by.username if payment.received_by else "—"),
        ("Date",            date.today().strftime('%d %b %Y')),
    ]
    _send_alert(
        subject=f"[ERP] Payment Received — Rs. {_int(payment.amount):,} ({buyer})",
        text_body=f"PAYMENT RECEIVED\nRs. {_int(payment.amount):,} via {payment.payment_method}",
        html_body=_html_wrap("Payment Received", rows, color="#1E40AF", note=""),
    )


# ─────────────────────────────────────────────────────────────────────────────
# 3. LOW STOCK ALERTS
# ─────────────────────────────────────────────────────────────────────────────

def check_chemical_stock_alerts():
    from decolorization.models import ChemicalStock
    critical = [
        c for c in ChemicalStock.objects.all()
        if float(c.total_stock or 0) > 0
        and (float(c.remaining_stock or 0) / float(c.total_stock)) < CHEMICAL_LOW_PCT
    ]
    if not critical:
        return
    rows = [(c.chemical_name,
             f"{_int(c.remaining_stock):,} / {_int(c.total_stock):,} {c.unit_of_measure} "
             f"({round(float(c.remaining_stock)/float(c.total_stock)*100,1)}%)")
            for c in critical]
    _send_alert(
        subject=f"[ERP] Low Chemical Stock — {len(critical)} item(s)",
        text_body="\n".join(f"{l}: {v}" for l, v in rows),
        html_body=_html_wrap(f"{len(critical)} Chemical(s) Critically Low", rows,
                             color="#DC2626", note="Restock required immediately."),
    )


def check_warehouse_stock_alerts():
    from warehouse.models import Stock
    total_kg = _live(Stock.objects.filter(status='Approved')).aggregate(
        t=Sum('unloading_weight'))['t'] or 0
    if float(total_kg) < WAREHOUSE_LOW_KG:
        rows = [
            ("Current Stock", f"{_int(total_kg):,} kg"),
            ("Threshold",     f"{WAREHOUSE_LOW_KG:,} kg"),
            ("Action",        "Restock required"),
        ]
        _send_alert(
            subject=f"[ERP] Warehouse Stock Low — {_int(total_kg):,} kg",
            text_body=f"Warehouse approved stock is {_int(total_kg):,} kg — below threshold.",
            html_body=_html_wrap("Warehouse Stock Running Low", rows,
                                 color="#D97706", note=""),
        )


# ─────────────────────────────────────────────────────────────────────────────
# 4. AUTO STATUS UPDATES
# ─────────────────────────────────────────────────────────────────────────────

def auto_update_order_payment_status(order):
    from sales.models import Payment
    total_price = float(order.total_price or 0)
    total_paid  = float(
        _live(Payment.objects.filter(sales_order=order))
        .aggregate(t=Sum('amount'))['t'] or 0
    )
    if total_price <= 0:
        return
    new_status = 'Paid' if total_paid >= total_price else ('Partial' if total_paid > 0 else 'Pending')
    if order.payment_status != new_status:
        order.payment_status = new_status
        order.save(update_fields=['payment_status'])
        logger.info(f"Order #{order.id} payment_status -> {new_status}")


def auto_update_fabric_status(fabric_stock_id):
    from sorting.models import FabricStock, SortingSession
    try:
        fabric = FabricStock.objects.get(id=fabric_stock_id)
    except FabricStock.DoesNotExist:
        return
    sessions = SortingSession.objects.filter(fabric=fabric)
    if not sessions.exists():
        new_status = 'In Warehouse'
    elif sessions.filter(status__in=['In Progress', 'Pending']).exists():
        new_status = 'In Sorting'
    elif sessions.filter(status='Completed').count() == sessions.count():
        new_status = 'Sorted'
    else:
        new_status = 'In Sorting'
    if fabric.status != new_status:
        fabric.status = new_status
        fabric.save(update_fields=['status'])


def auto_update_tank_status(tank_id):
    from decolorization.models import Tank, DecolorizationSession
    try:
        tank = Tank.objects.get(id=tank_id)
    except Tank.DoesNotExist:
        return
    sessions   = DecolorizationSession.objects.filter(tank=tank)
    fabric_qty = float(tank.fabric_quantity or 0)
    if sessions.filter(status='In Progress').exists():
        new_status = 'Processing'
    elif sessions.filter(status='Completed').exists():
        new_status = 'Completed'
    elif fabric_qty > 0:
        new_status = 'Filled'
    else:
        new_status = 'Empty'
    if tank.tank_status != new_status:
        tank.tank_status = new_status
        tank.save(update_fields=['tank_status'])


# ─────────────────────────────────────────────────────────────────────────────
# 5A. Build Excel — Monthly Sales
# ─────────────────────────────────────────────────────────────────────────────

def _build_monthly_excel(month_name, year, order_qs, pay_qs, order_agg, pay_agg):
    """
    Build a multi-sheet monthly sales Excel workbook.
    Returns raw bytes ready to attach to an email.
    """
    wb  = openpyxl.Workbook()
    rev = _int(order_agg['revenue'])
    col = _int(pay_agg['total'])

    # ── Sheet 1: Summary ─────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Summary"
    _xl_title(ws,
              "Textile ERP — Recycling System",
              f"Monthly Sales Report — {month_name} {year}",
              f"Generated: {date.today().strftime('%d/%m/%Y')}", 4)

    ws.cell(5, 1, "MONTHLY SALES SUMMARY").font = Font(bold=True, color=_BLUE_HEX, size=12)
    _xl_summary(ws, 6, 1, [
        ("Period",                   f"{month_name} {year}"),
        ("Total Orders",             str(order_agg['count'] or 0)),
        ("Total Revenue (Rs.)",      f"{rev:,}"),
        ("Amount Collected (Rs.)",   f"{col:,}"),
        ("Pending Amount (Rs.)",     f"{max(0, rev - col):,}"),
        ("Total Weight Sold (kg)",   f"{_int(order_agg['weight']):,}"),
        ("Avg Price / kg (Rs.)",     f"{_int(order_agg.get('avg_price', 0)):,}"),
    ])
    _cw(ws, 1, 30); _cw(ws, 2, 22)

    # ── Sheet 2: All Orders ───────────────────────────────────────────────────
    ws2 = wb.create_sheet("Sales Orders")
    _xl_title(ws2, "Textile ERP — Recycling System",
              f"Sales Orders — {month_name} {year}", "", 9)
    _xl_header(ws2, 5, ["#", "Order ID", "Buyer", "Contact",
                         "Quality", "Weight (kg)", "Price/kg (Rs.)",
                         "Total (Rs.)", "Payment"])
    for i, o in enumerate(order_qs, 1):
        _xl_row(ws2, 5 + i, [
            i, f"#{o.id}", o.buyer_name, o.buyer_contact or "—",
            o.fabric_quality or "—",
            _int(o.weight_sold), _int(o.price_per_kg), _int(o.total_price),
            o.payment_status,
        ], money_cols={7, 8}, alt=(i % 2 == 0))
    for c, w in zip(range(1, 10), [4, 9, 26, 16, 12, 12, 14, 14, 12]):
        _cw(ws2, c, w)

    # ── Sheet 3: Payments ─────────────────────────────────────────────────────
    ws3 = wb.create_sheet("Payments")
    _xl_title(ws3, "Textile ERP — Recycling System",
              f"Payments — {month_name} {year}", "", 6)
    _xl_header(ws3, 5, ["#", "Order", "Buyer", "Amount (Rs.)", "Method", "Reference"])
    for i, p in enumerate(pay_qs, 1):
        buyer = p.sales_order.buyer_name if p.sales_order else "—"
        _xl_row(ws3, 5 + i, [
            i, f"#{p.sales_order_id}", buyer,
            _int(p.amount), p.payment_method, p.reference_number or "—",
        ], money_cols={4}, alt=(i % 2 == 0))
    for c, w in zip(range(1, 7), [4, 9, 26, 16, 16, 18]):
        _cw(ws3, c, w)

    return _wb_bytes(wb)


# ─────────────────────────────────────────────────────────────────────────────
# 5B. Build Excel — Daily Production
# ─────────────────────────────────────────────────────────────────────────────

def _build_daily_excel(today, stock_count, stock_wt,
                       si, so, sw, di, do_, dw,
                       sort_sessions_qs, decolor_sessions_qs):
    """
    Build a multi-sheet daily production Excel workbook.
    Returns raw bytes ready to attach to an email.
    """
    wb = openpyxl.Workbook()
    date_str = today.strftime('%d %B %Y')

    # ── Sheet 1: Summary ─────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Summary"
    _xl_title(ws,
              "Textile ERP — Recycling System",
              f"Daily Production Report — {today.strftime('%A, %d %B %Y')}",
              f"Generated: {date.today().strftime('%d/%m/%Y')}", 4)

    ws.cell(5, 1, "PRODUCTION SUMMARY").font = Font(bold=True, color=_BLUE_HEX, size=12)
    _xl_summary(ws, 6, 1, [
        ("Date",                         today.strftime('%A, %d %B %Y')),
        ("Warehouse Stock Entries",       str(stock_count)),
        ("Warehouse Weight Received (kg)", f"{_int(stock_wt):,}"),
        ("Sorting Sessions",              str(sort_sessions_qs.count())),
        ("Sorting Input (kg)",            f"{si:,}"),
        ("Sorting Output (kg)",           f"{so:,}"),
        ("Sorting Waste (kg)",            f"{sw:,}"),
        ("Sorting Efficiency",            f"{round(so/si*100,1) if si else 0}%"),
        ("Decolorization Sessions",       str(decolor_sessions_qs.count())),
        ("Decolorization Input (kg)",     f"{di:,}"),
        ("Decolorization Output (kg)",    f"{do_:,}"),
        ("Decolorization Waste (kg)",     f"{dw:,}"),
        ("Decolorization Efficiency",     f"{round(do_/di*100,1) if di else 0}%"),
    ])
    _cw(ws, 1, 34); _cw(ws, 2, 22)

    # ── Sheet 2: Sorting Sessions ─────────────────────────────────────────────
    ws2 = wb.create_sheet("Sorting Sessions")
    _xl_title(ws2, "Textile ERP — Recycling System",
              f"Sorting Sessions — {date_str}", "", 7)
    _xl_header(ws2, 5, ["#", "Fabric", "Supervisor", "Unit",
                         "Input (kg)", "Output (kg)", "Waste (kg)"])
    for i, s in enumerate(sort_sessions_qs, 1):
        _xl_row(ws2, 5 + i, [
            i,
            s.fabric.material_type if s.fabric else "—",
            s.supervisor.username  if s.supervisor else "—",
            s.unit or "—",
            _int(s.quantity_taken),
            _int(s.quantity_sorted),
            _int(s.waste_quantity),
        ], money_cols={5, 6, 7}, alt=(i % 2 == 0))
    for c, w in zip(range(1, 8), [4, 28, 16, 10, 12, 12, 12]):
        _cw(ws2, c, w)

    # ── Sheet 3: Decolorization Sessions ─────────────────────────────────────
    ws3 = wb.create_sheet("Decolorization Sessions")
    _xl_title(ws3, "Textile ERP — Recycling System",
              f"Decolorization Sessions — {date_str}", "", 7)
    _xl_header(ws3, 5, ["#", "Tank", "Fabric", "Supervisor",
                         "Input (kg)", "Output (kg)", "Waste (kg)"])
    for i, s in enumerate(decolor_sessions_qs, 1):
        _xl_row(ws3, 5 + i, [
            i,
            s.tank.name            if s.tank else "—",
            s.fabric.material_type if s.fabric else "—",
            s.supervisor.username  if s.supervisor else "—",
            _int(s.input_quantity),
            _int(s.output_quantity),
            _int(s.waste_quantity),
        ], money_cols={5, 6, 7}, alt=(i % 2 == 0))
    for c, w in zip(range(1, 8), [4, 14, 28, 16, 12, 12, 12]):
        _cw(ws3, c, w)

    return _wb_bytes(wb)


# ─────────────────────────────────────────────────────────────────────────────
# 6. SCHEDULED REPORT FUNCTIONS
# ─────────────────────────────────────────────────────────────────────────────

def send_monthly_sales_summary():
    """
    Sends monthly sales HTML email WITH Excel attachment to management.
    Covers the previous calendar month automatically.

    Run manually:  python manage.py send_monthly_report
    Windows Task Scheduler: see task_scheduler_monthly.xml
    """
    from sales.models import SalesOrder, Payment

    today            = date.today()
    first_of_this    = today.replace(day=1)
    last_month_end   = first_of_this - timedelta(days=1)
    last_month_start = last_month_end.replace(day=1)
    month_name       = calendar.month_name[last_month_end.month]
    year             = last_month_end.year

    start = timezone.make_aware(
        timezone.datetime.combine(last_month_start, timezone.datetime.min.time())
    )
    end = timezone.make_aware(
        timezone.datetime.combine(last_month_end, timezone.datetime.max.time())
    )

    order_qs  = _live(_date_range(SalesOrder.objects.all(), start, end)).select_related('fabric')
    pay_qs    = _live(_date_range(Payment.objects.all(),    start, end)).select_related('sales_order')

    order_agg = order_qs.aggregate(
        count=Count('id'), revenue=Sum('total_price'),
        weight=Sum('weight_sold'), avg_price=Avg('price_per_kg'),
    )
    pay_agg   = pay_qs.aggregate(total=Sum('amount'))

    rev = _int(order_agg['revenue'])
    col = _int(pay_agg['total'])

    # ── HTML email body ───────────────────────────────────────────────────────
    rows = [
        ("Period",              f"{month_name} {year}"),
        ("Total Orders",        str(order_agg['count'] or 0)),
        ("Total Revenue",       f"Rs. {rev:,}"),
        ("Amount Collected",    f"Rs. {col:,}"),
        ("Pending Amount",      f"Rs. {max(0, rev - col):,}"),
        ("Total Weight Sold",   f"{_int(order_agg['weight']):,} kg"),
        ("Avg Price / kg",      f"Rs. {_int(order_agg.get('avg_price', 0)):,}"),
    ]
    html = _html_wrap(
        f"Monthly Sales Report — {month_name} {year}", rows,
        color="#059669",
        note="Full Excel report attached below.",
    )
    text = "\n".join(f"{k}: {v}" for k, v in rows)

    # ── Build Excel attachment ────────────────────────────────────────────────
    xl_bytes    = _build_monthly_excel(
        month_name, year, order_qs, pay_qs, order_agg, pay_agg
    )
    xl_filename = f"Monthly_Sales_{year}_{last_month_end.month:02d}_{month_name}.xlsx"

    # ── Send with attachment ──────────────────────────────────────────────────
    try:
        _send_with_attachment(
            subject=f"[ERP] Monthly Sales Report — {month_name} {year}",
            text_body=text,
            html_body=html,
            xl_bytes=xl_bytes,
            xl_filename=xl_filename,
        )
    except Exception as exc:
        logger.error(f"send_monthly_sales_summary failed: {exc}")
        raise


def send_daily_production_summary():
    """
    Sends daily production HTML email WITH Excel attachment to management.

    Run manually:  python manage.py send_daily_report
    Windows Task Scheduler: see task_scheduler_daily.xml
    """
    from warehouse.models import Stock
    from sorting.models import SortingSession
    from decolorization.models import DecolorizationSession

    today = date.today()
    start = timezone.make_aware(
        timezone.datetime.combine(today, timezone.datetime.min.time())
    )
    end = start + timedelta(days=1)

    # Warehouse
    stock_qs    = _live(_date_range(Stock.objects.all(), start, end - timedelta(seconds=1)))
    stock_count = stock_qs.count()
    stock_wt    = stock_qs.aggregate(t=Sum('unloading_weight'))['t'] or 0

    # Sorting — use _date_range() which auto-detects the right field (start_date)
    sort_qs  = _date_range(
        SortingSession.objects.all(), start, end - timedelta(seconds=1)
    ).select_related('fabric', 'supervisor')
    sort_agg = sort_qs.aggregate(
        inp=Sum('quantity_taken'),
        out=Sum('quantity_sorted'),
        wst=Sum('waste_quantity'),
    )

    # Decolorization — use _date_range() which auto-detects the right field (start_date)
    dec_qs  = _date_range(
        DecolorizationSession.objects.all(), start, end - timedelta(seconds=1)
    ).select_related('tank', 'fabric', 'supervisor')
    dec_agg = dec_qs.aggregate(
        inp=Sum('input_quantity'),
        out=Sum('output_quantity'),
        wst=Sum('waste_quantity'),
    )

    si  = _int(sort_agg['inp'])
    so  = _int(sort_agg['out'])
    sw  = _int(sort_agg['wst'])
    di  = _int(dec_agg['inp'])
    do_ = _int(dec_agg['out'])
    dw  = _int(dec_agg['wst'])

    # ── HTML email body ───────────────────────────────────────────────────────
    rows = [
        ("Date",                    today.strftime('%A, %d %B %Y')),
        ("Warehouse Receipts",      f"{stock_count} entries ({_int(stock_wt):,} kg)"),
        ("Sorting Input",           f"{si:,} kg"),
        ("Sorting Output",          f"{so:,} kg"),
        ("Sorting Waste",           f"{sw:,} kg ({round(sw/si*100,1) if si else 0}%)"),
        ("Decolorization Input",    f"{di:,} kg"),
        ("Decolorization Output",   f"{do_:,} kg"),
        ("Decolorization Waste",    f"{dw:,} kg ({round(dw/di*100,1) if di else 0}%)"),
    ]
    html = _html_wrap(
        f"Daily Production Report — {today.strftime('%d %b %Y')}", rows,
        color="#1E40AF",
        note="Full Excel report attached below.",
    )
    text = "\n".join(f"{k}: {v}" for k, v in rows)

    # ── Build Excel attachment ────────────────────────────────────────────────
    xl_bytes    = _build_daily_excel(
        today, stock_count, stock_wt,
        si, so, sw, di, do_, dw,
        sort_qs, dec_qs,
    )
    xl_filename = f"Daily_Production_{today.strftime('%Y-%m-%d')}.xlsx"

    # ── Send with attachment ──────────────────────────────────────────────────
    try:
        _send_with_attachment(
            subject=f"[ERP] Daily Production Report — {today.strftime('%d %b %Y')}",
            text_body=text,
            html_body=html,
            xl_bytes=xl_bytes,
            xl_filename=xl_filename,
        )
    except Exception as exc:
        logger.error(f"send_daily_production_summary failed: {exc}")
        raise