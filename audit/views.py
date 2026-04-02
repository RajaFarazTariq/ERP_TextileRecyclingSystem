# audit/views.py
"""
REST API for querying audit logs.
Endpoints:
  GET /api/audit/logs/                       → paginated list (admin only)
  GET /api/audit/logs/?model=SalesOrder      → filter by model
  GET /api/audit/logs/?user=3                → filter by user ID
  GET /api/audit/logs/?action=UPDATE         → filter by action
  GET /api/audit/logs/?start=YYYY-MM-DD      → filter by date range
  GET /api/audit/logs/<id>/                  → single entry detail
  GET /api/audit/logs/export/                → Excel export
"""

import io
from datetime import date, timedelta

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from django.http import HttpResponse
from rest_framework import serializers, viewsets, filters
from rest_framework.decorators import api_view, permission_classes, action
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.pagination import PageNumberPagination

from .models import AuditLog


# ── Serializer ────────────────────────────────────────────────────────────────

class AuditLogSerializer(serializers.ModelSerializer):
    timestamp_display = serializers.SerializerMethodField()

    def get_timestamp_display(self, obj):
        return obj.timestamp.strftime('%d %b %Y, %H:%M:%S')

    class Meta:
        model  = AuditLog
        fields = [
            'id', 'username', 'user_role', 'action', 'model_name',
            'object_id', 'object_repr', 'changes',
            'ip_address', 'endpoint', 'timestamp', 'timestamp_display',
        ]


# ── Pagination ────────────────────────────────────────────────────────────────

class AuditPagination(PageNumberPagination):
    page_size            = 50
    page_size_query_param = 'page_size'
    max_page_size        = 200


# ── ViewSet ───────────────────────────────────────────────────────────────────

class AuditLogViewSet(viewsets.ReadOnlyModelViewSet):
    """
    Read-only. Admin-only access.
    Supports filtering by: model, user, action, start, end, search.
    """
    serializer_class   = AuditLogSerializer
    permission_classes = [IsAuthenticated]
    pagination_class   = AuditPagination
    filter_backends    = [filters.SearchFilter, filters.OrderingFilter]
    search_fields      = ['username', 'model_name', 'object_repr', 'action']
    ordering_fields    = ['timestamp', 'action', 'model_name']
    ordering           = ['-timestamp']

    def get_queryset(self):
        # Only admin can see all logs
        user = self.request.user
        qs   = AuditLog.objects.all()

        if not (user.is_staff or getattr(user, 'role', '') == 'admin'):
            # Non-admins can only see their own actions
            qs = qs.filter(user=user)

        # Query param filters
        model  = self.request.query_params.get('model')
        uid    = self.request.query_params.get('user')
        action = self.request.query_params.get('action')
        start  = self.request.query_params.get('start')
        end    = self.request.query_params.get('end')

        if model:
            qs = qs.filter(model_name__iexact=model)
        if uid:
            qs = qs.filter(user_id=uid)
        if action:
            qs = qs.filter(action__iexact=action)
        if start:
            try:
                qs = qs.filter(timestamp__date__gte=date.fromisoformat(start))
            except ValueError:
                pass
        if end:
            try:
                qs = qs.filter(timestamp__date__lte=date.fromisoformat(end))
            except ValueError:
                pass

        return qs

    @action(detail=False, methods=['get'], url_path='export')
    def export_excel(self, request):
        """Export audit logs to Excel."""
        qs = self.get_queryset()[:500]   # cap at 500 rows for export

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Audit Logs"

        thin = Side(style="thin", color="CBD5E1")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        blue   = PatternFill("solid", fgColor="1E40AF")
        alt    = PatternFill("solid", fgColor="F8FAFC")
        white  = PatternFill("solid", fgColor="FFFFFF")

        # Title
        ws.merge_cells("A1:H1")
        tc = ws.cell(1, 1, "Textile ERP — Audit Log Export")
        tc.font = Font(bold=True, size=14, color="1E40AF")
        tc.alignment = Alignment(horizontal="center")
        ws.merge_cells("A2:H2")
        dc = ws.cell(2, 1, f"Generated: {date.today()} | Records: {len(qs)}")
        dc.font = Font(size=9, italic=True, color="64748B")
        dc.alignment = Alignment(horizontal="center")

        headers = ["#", "Timestamp", "User", "Role", "Action", "Model", "Record ID", "Description"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(4, col, h)
            cell.font = Font(bold=True, color="FFFFFF", size=10)
            cell.fill = blue
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for i, log in enumerate(qs, 1):
            fill = alt if i % 2 == 0 else white
            row_data = [
                i,
                log.timestamp.strftime('%Y-%m-%d %H:%M:%S'),
                log.username,
                log.user_role,
                log.action,
                log.model_name,
                log.object_id,
                log.object_repr[:80],
            ]
            for col, val in enumerate(row_data, 1):
                cell = ws.cell(4 + i, col, val)
                cell.font = Font(size=10, color="334155")
                cell.fill = fill
                cell.border = border

        for col, w in zip(range(1, 9), [5, 22, 16, 14, 12, 18, 10, 40]):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        response = HttpResponse(
            buf.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="audit_log_{date.today()}.xlsx"'
        return response


# ── Summary stats endpoint ────────────────────────────────────────────────────

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def audit_summary(request):
    """Returns quick stats for the audit dashboard."""
    from django.db.models import Count
    user = request.user

    if not (user.is_staff or getattr(user, 'role', '') == 'admin'):
        return Response({'error': 'Admin access required.'}, status=403)

    today     = date.today()
    week_ago  = today - timedelta(days=7)
    month_ago = today - timedelta(days=30)

    qs = AuditLog.objects.all()

    return Response({
        'total_logs':      qs.count(),
        'today':           qs.filter(timestamp__date=today).count(),
        'this_week':       qs.filter(timestamp__date__gte=week_ago).count(),
        'this_month':      qs.filter(timestamp__date__gte=month_ago).count(),
        'by_action':       list(qs.values('action').annotate(count=Count('id')).order_by('-count')),
        'by_model':        list(qs.values('model_name').annotate(count=Count('id')).order_by('-count')[:10]),
        'recent_activity': list(
            qs.values('username', 'action', 'model_name', 'object_repr', 'timestamp')
              .order_by('-timestamp')[:10]
        ),
    })
